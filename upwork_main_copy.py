
from openai import OpenAI
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException, ElementClickInterceptedException
from screeninfo import get_monitors
import undetected_chromedriver as uc
import openpyxl as pyxl
import time
import random
import math
import re
import requests
import pickle
import os
import threading
import inspect
from dotenv import load_dotenv

import gspread
from gspread_formatting import *
from oauth2client.service_account import ServiceAccountCredentials
import json
from enum import Enum

def is_json(variable):
    try:
        json.loads(variable)
    except (TypeError, json.JSONDecodeError):
        return False
    return True

# Load the .env file
load_dotenv()

# Now you can access the API key using os.getenv
api_key = "Ask for API key"

#Global Variables
max_wait_time = 60
number_of_suppliers_to_contact = 2
max_retries = 5
chat_product_dict = {} #Key: supplier name, Value: list queue of product tuples (first is current product chat)
chat_step_dict = {} #Key: supplier name, Value: dictionary key: index, value is the chat questions left
chat_product_lock = threading.Lock()
chat_step_lock = threading.Lock()
index_step_dict_lock = threading.Lock()
excel_lock = threading.Lock()
chat_dict_loc = "chat_product_dict.pkl"
chat_step_dict_loc = "chat_step_dict.pkl"



def with_cooldown(func, max_attempts=5, initial_wait=3):
    attempt = 0
    wait_time = initial_wait

    while attempt < max_attempts:
        try:
            time.sleep(wait_time)
            result = func()
            if isinstance(result, dict) and result.get('error', '') == 'Rate limit exceeded':
                print(f"Rate limit exceeded. Waiting {wait_time} seconds before retrying...")
                wait_time *= 2  # Exponential back-off
                attempt += 1
            else:
                return result
        except Exception as e:
            print(f"An error occurred: {e}")
            if "Rate limit exceeded" in str(e):
                print(f"Rate limit exceeded in exception. Waiting {wait_time} seconds before retrying...")
                wait_time *= 2
                attempt += 1
            else:
                return None  # Return None immediately on other exceptions

    print("Maximum retry attempts reached.")
    return None

def extract_asin(input_string):
    # Regular expression to match an Amazon ASIN within various contexts
    match = re.search(r'(?:dp|ASIN|gp/product)/([A-Z0-9]{10})', input_string)
    if match:
        return match.group(1)  # Return the ASIN
    else:
        return "ASIN not found"
    
def clear_chat_dicts():
    global chat_product_dict
    global chat_product_lock
    global chat_step_dict
    global chat_step_lock

    with chat_product_lock:
        chat_product_dict = {}
    with chat_step_lock:
        chat_step_dict = {}

    write_dict_to_file(chat_dict_loc, chat_product_dict, chat_product_lock)
    write_dict_to_file(chat_step_dict_loc, chat_step_dict, chat_step_lock)

def read_chat_dicts(specific_dict=None):
    global chat_product_dict
    global chat_product_lock
    global chat_step_dict
    global chat_step_lock

    if specific_dict == None:
        with chat_product_lock:
            if os.path.exists(chat_dict_loc):
                with open(chat_dict_loc, 'rb') as chat_file:
                    chat_product_dict = pickle.load(chat_file)
        with chat_step_lock:
            if os.path.exists(chat_step_dict_loc):
                with open(chat_step_dict_loc, 'rb') as chat_file:
                    chat_step_dict = pickle.load(chat_file)
    else:
        if specific_dict == "chat_product_dict":
            with chat_product_lock:
                if os.path.exists(chat_dict_loc):
                    with open(chat_dict_loc, 'rb') as chat_file:
                        chat_product_dict = pickle.load(chat_file)
        elif specific_dict == "chat_step_dict":
            with chat_step_lock:
                if os.path.exists(chat_step_dict_loc):
                    with open(chat_step_dict_loc, 'rb') as chat_file:
                        chat_step_dict = pickle.load(chat_file)


def write_dict_to_file(file_path, global_dict, global_lock):
    
    with global_lock:
        write_retries = 0
        while True:
            with open(file_path, 'wb') as file:
                pickle.dump(global_dict, file)

            # Attempt to read back the dictionary to verify its integrity
            try:
                with open(file_path, 'rb') as file:
                    data_read_back = pickle.load(file)
                # Compare the original dictionary with the read-back data
                if data_read_back == global_dict:
                    break
                else:
                    write_retries += 1
                    if write_retries > max_retries:
                        raise ValueError(f"Pickle file for {file_path} corrupted after writing.") 
            except (pickle.PickleError, EOFError) as e:
                write_retries += 1
                if write_retries > max_retries:
                    raise ValueError(f"Pickle file for {file_path} corrupted after writing.")


def random_sleep(min_time, max_time):
    wait_time = random.uniform(min_time, max_time)
    time.sleep(wait_time)

def query_openai(prompt, model, max_retries=max_retries):
    base_wait = 1  # Base wait time in seconds

    #Obtain OpenAI API Access
    client = OpenAI(api_key=api_key)

    for i in range(max_retries):
        try:
            if model == "titles":
                #Get simplified titles
                response = client.chat.completions.create(
                # model="ft:gpt-3.5-turbo-0125:personal:simplified-titles:96s6KQpr",
                model="gpt-4o-2024-05-13",
                messages=[
                    {"role": "system", "content": "You are a program that needs to return a list of simplified titles from an 'Initial title' and must abide by the given rules. Rule 1: Begin with the 'Initial title' and with each iteration make the simplified title generated less detailed (less characters). Rule 2: In the list of simplified titles returned, have the simplified titles ordered from 1) most detailed (most characters) to 10) least detailed (least characters). Rule 3: Only return simplified titles that are less than 50 characters in total length. Rule 4: Return 10 simplified titles. Rule 5: Do not include character count in the simplified titles. Rule 6: Exclude any brand names from the returned simplified titles unless the product is specific for a brand such as 'brand replacement parts'."},
                    {"role": "user", "content": f"{prompt}"}
                ]
                )
            elif model == "analysing":

                current_questions, supplier_response = prompt
                print(f'[175] f = {current_questions}')
                print(f'[176] s = {supplier_response}')

                messages=[
                    {"role": "system", "content": "The input is answer of supplier about it. First, Analyse supplier's answer and seventh answer must be set 'picture' if the supplier needs a picture of the product." + " basic questions are choosen among 7 as below. 1. Are you selling {product_name}? 2. What is the EXW price for {quantity} units? 3. Can I get a sample? 4. What are the package dimensions? 5. What is the package weight for {quantity} units? 6. Does the product come unbranded? 7. Would I be able to get a picture? There can be each different content in product_name, quantity. The result should be following only the JSON Style. output order should be same like order of inputed basic question. The element count of JSON structure of output style follow basic question count. Below is the case where there are 7 questions in the basic question. Output style: {1: value1, 2: value2, 3: value3, 4: value4, 5: value5, 6: value6, 7: value7} note: - Indexes of Output must always start from number 1, not string. - value1 should be 'yes', 'no' or 'unsure' string. - value2 should be only number value or 'unsure' string if a exact numerical value is provided, the answer is maximum of those values. - value3 should be 'yes', 'no' or 'unsure' string. - value4 should be general sentence or 'unsure' string if answer of suppiler is placeholder or not a exact numerical value, the answer is 'unsure' string. - value5 should be general sentence or 'unsure' string if answer of suppiler is placeholder or not a exact numerical value, the answer is 'unsure' string. - value6 should be 'yes', 'no' or 'unsure' string. - value7 should be 'yes', 'no', 'picture' or 'unsure' string."},
                    {"role": "user", "content": supplier_response}
                ]

                response = client.chat.completions.create(
                    model="gpt-4o-2024-05-13",
                    messages=messages,
                    response_format={"type": "json_object"}
                )
                print(f"[188] response = {response}")
            return response.choices[0].message.content
        except OpenAI.RateLimitError:
            wait_time = base_wait * math.pow(2, i)  # Exponential backoff formula
            print(f"Rate limit exceeded, waiting for {wait_time} seconds before retrying...")
            time.sleep(wait_time)
        except OpenAI.OpenAIError as e:
            print(f"An OpenAI API error occurred: {e}")
            break  # Exit the loop for non-rate-limit errors
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            break  # Exit the loop for other errors
    return None  # Return None if all retries fail

def initialize_alibaba_search():
    global max_wait_time

    # Open the Alibaba
    # proxy = '38.154.126.69:8800'
    # chrome_options = Options()
    # chrome_options.add_argument(f'--proxy-server={proxy}')
    # chrome_options.add_argument("--disable-notifications")
    # driver = uc.Chrome(options=chrome_options)
    driver = uc.Chrome()
    monitor = get_monitors()[0]
    screenHeight = monitor.height
    screenWidth = monitor.width
    wait = WebDriverWait(driver, max_wait_time)
    driver.set_window_size(screenWidth * 0.75, screenHeight)
    # driver.set_window_position((-1) * screenWidth, 0)
    driver.get('https://www.alibaba.com/')

    # Load and add cookies from the file
    if os.path.exists("alibaba_login_cookies.pkl"):
        with open("alibaba_login_cookies.pkl", "rb") as cookies_file:
            print("Loading cookies...")
            cookies = pickle.load(cookies_file)
            for cookie in cookies:
                driver.add_cookie(cookie)
        
        #Reload View
        random_sleep(0, 1)
        driver.get('https://www.alibaba.com/')
        print("Reload view with cookies")
    else:
        #Sign in
        try:
            driver.get('https://login.alibaba.com/newlogin/icbuLogin.htm?return_url=https%3A%2F%2Fwww.alibaba.com%2F&_lang=')
        except Exception as e:
            raise ValueError(f"Sign in button not found, exception: {e}")

        #Click login with email
        random_sleep(0, 2)
        try:
            email = 'email'
            password = 'pass'
            login_with_email_css = '.sif_form.sif_form-account'
            password_css = '.sif_form.sif_form-password'
            submit_btn_xpath = "//button[contains(@class, 'sif_form-submit')]"

            login_text_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, login_with_email_css)))
            password_text_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, password_css)))
            submit_btn = wait.until(EC.element_to_be_clickable((By.XPATH, submit_btn_xpath)))

            login_text_box.click()
            login_text_box.send_keys(email)
            password_text_box.send_keys(password)
            submit_btn.click()

        except Exception as e:
            raise ValueError(f"Login text box not found, exception: {e}")
        
        random_sleep(0, 2)
        cookies = driver.get_cookies()
        with open("alibaba_login_cookies.pkl", "wb") as cookies_file:
            pickle.dump(cookies, cookies_file)
        print("Cookies saved to 'alibaba_login_cookies.pkl'")

    return driver, wait

def final_input_interaction(driver, wait, description_tuple, supplier_name):

    #Get product info
    quantity, rfq_product_name, image_url, max_exw_price, max_size = description_tuple

    #Input RFQ Quantity
    random_sleep(0, 1)
    rfq_quantity_input_xpath = "//span[contains(@class, 'next-input') and contains(@class, 'next-small') and contains(@class, 'next-noborder')]/input"
    try:
        rfq_quantity_input = driver.find_element(By.XPATH, rfq_quantity_input_xpath)
        rfq_quantity_input.click()
        random_sleep(0, 1)
        driver.execute_script("arguments[0].value = '';", rfq_quantity_input)
        random_sleep(0, 1)
        rfq_quantity_input.send_keys(quantity)
    except Exception as e:
        rfq_quantity_input_xpath = "//div[contains(@class, 'quantity-wrap')]/input"
        try:
            rfq_quantity_input = driver.find_element(By.XPATH, rfq_quantity_input_xpath)
            rfq_quantity_input.click()
            random_sleep(0, 1)
            driver.execute_script("arguments[0].value = '';", rfq_quantity_input)
            random_sleep(0, 1)
            rfq_quantity_input.send_keys(quantity)
        except Exception as e:
            print(f"No RFQ quantity input found when contacting supplier, exception: {e}")
            return True
    
    #Input RFQ Inquiry
    random_sleep(0, 1)
    inquiry_input_id = "inquiry-content"
    inquiry_message = f"Hello, I am interested in purchasing the following product: {rfq_product_name}. Could you please provide me with a quote?"
    try:
        inquiry_input = driver.find_element(By.ID, inquiry_input_id)
        inquiry_input.click()
        random_sleep(0, 1)
        inquiry_input.send_keys(inquiry_message)
    except Exception as e:
        inquiry_input_xpath = "//textarea[contains(@class, 'content-input')]"
        try:
            inquiry_input = driver.find_element(By.XPATH, inquiry_input_xpath)
            inquiry_input.click()
            random_sleep(0, 1)
            inquiry_input.send_keys(inquiry_message)
        except Exception as e:
            print(f"No inquiry input found when contacting supplier, exception: {e}")
            return True
    
    #Download Image to send
    image_download_iter = 0
    current_dir = os.path.dirname(os.path.abspath(__file__))
    tmp_file_path = os.path.join(current_dir, "tmp_image.jpg")
    while True:
        try:
            random_sleep(1, 2)
            image_response = requests.get(image_url, timeout=10)  # Add timeout
            if image_response.status_code == 200:
                with open(tmp_file_path, 'wb') as tmp_file:
                    tmp_file.write(image_response.content)
                break
            else:
                print(f"Error: Received status code {image_response.status_code} for URL {image_url}")
        except requests.exceptions.RequestException as e:
            print(f"Attempt {image_download_iter + 1} failed: {e}")
        if image_download_iter > max_retries:
            print(f"Failed to download image after {max_retries} attempts")
            return True
        image_download_iter += 1

    #Input RFQ Image
    random_sleep(0, 1)
    image_input_id = "ksu-fileserver-1"
    try:
        image_input = driver.find_element(By.ID, image_input_id)
        driver.execute_script("arguments[0].scrollIntoView();", image_input)
        image_input.send_keys(tmp_file_path)
    except Exception as e:
        image_input_xpath = "/html/body/div[1]/div/div/div/div[2]/div[2]/div/div[1]/input"
        try:
            image_input = driver.find_element(By.XPATH, image_input_xpath)
            driver.execute_script("arguments[0].scrollIntoView();", image_input)
            image_input.send_keys(tmp_file_path)
        except Exception as e:
            print(f"Failed to send image when contacting supplier, exception: {e}")
            return True
        
    #Send Inquiry
    random_sleep(0, 1)
    send_inquiry_button_xpath = "/html/body/div[1]/div/div/div/div[3]/button"
    try:
        send_inquiry_button = driver.find_element(By.XPATH, send_inquiry_button_xpath)
        send_inquiry_button.click()
    except Exception as e:
        send_inquiry_button_xpath = "/html/body/div[2]/div/form/div[1]/div[2]/div/div/div[2]/input"
        try:
            send_inquiry_button = driver.find_element(By.XPATH, send_inquiry_button_xpath)
            send_inquiry_button.click()
            send_inquiry_button.click()
        except Exception as e:
            print(f"Failed to send inquiry for supplier, exception: {e}")
            return True

    #See if successful and delete temp file
    random_sleep(3, 4)
    succeeded_page_id = "alitalk-dialog-inquiry-succeed"
    try:
        driver.switch_to.default_content()
        driver.find_element(By.ID, succeeded_page_id)
        if os.path.exists(tmp_file_path):
            os.remove(tmp_file_path)
    except Exception as e:
        succeeded_page_icon_xpath = "//i[contains(@class, 'ui2-icon-success')]"
        try:
            driver.find_element(By.XPATH, succeeded_page_icon_xpath)
            if os.path.exists(tmp_file_path):
                os.remove(tmp_file_path)
        except Exception as e:
            print(f"Failed to confirm successful inquiry for supplier, exception: {e}")
            return True

    #Add chat product history (initial chat not included in chat since should be included in steps)
    with chat_product_lock:
        print(f"Adding to product dict for supplier : {supplier_name}")
        chat_product_dict[supplier_name] = []
        chat_product_dict[supplier_name].append(description_tuple)

    async_thread = threading.Thread(target=write_dict_to_file, args=(chat_dict_loc, chat_product_dict, chat_product_lock))
    async_thread.start()

    #Close inquiry tab
    random_sleep(0, 1)
    change_win_iter = 0
    while True:
        random_sleep(1, 2)
        try:
            if len(driver.window_handles) > 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            break
        except Exception as e:
            if change_win_iter > max_retries:
                print(f"Failed to switch windows after contacting supplier, exception: {e}")
                return True
            change_win_iter += 1
    
    return False

def send_initial_message(driver, wait, search_term, rfq_info, first_search, supplier_set):

    #Initialize variables
    global number_of_suppliers_to_contact
    global max_retries
    global chat_product_dict
    global chat_product_lock

    if first_search:
        #Search box
        try:
            search_bar_class = 'search-bar-placeholder'
            search_bar = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, search_bar_class)))
            search_bar.click()
        except Exception as e:
            pass
        
        random_sleep(0, 1)
        
    try:
        search_input_class = 'search-bar-input'
        search_input_box = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, search_input_class)))
        search_input_box.clear()
        random_sleep(0, 1)
        search_input_box.send_keys(search_term)
        random_sleep(1, 2)
        search_input_box.send_keys(Keys.RETURN)
    except Exception as e:
        print(f"Search input box not found, exception: {e}")
        return supplier_set
    
    random_sleep(2, 3)


    supplier_name_class = "search-card-e-company"
    retry_attempts = 0

    while retry_attempts < max_retries:
        try:
            text_elements = wait.until(EC.visibility_of_all_elements_located((By.CLASS_NAME, supplier_name_class)))
            break  # Exit loop if successful
        except TimeoutException as e:
            retry_attempts += 1
            print(f"Attempt {retry_attempts} failed: {e}")
            if retry_attempts >= max_retries:
                # Print page source for debugging
                page_source = driver.page_source
                with open("debug_page_source.html", "w", encoding="utf-8") as f:
                    f.write(page_source)
                return supplier_set
                # raise ValueError(f"No supplier names found for search term: {search_term}, exception: {e}") 
            time.sleep(2 ** retry_attempts)  # Exponential backoff

    if not text_elements:
        raise ValueError(f"No text elements found for search term: {search_term}")

    random_sleep(0, 1)
    supplier_image_class = "search-card-e-slider__wrapper"
    try:
        image_elements = wait.until(EC.visibility_of_all_elements_located((By.CLASS_NAME, supplier_image_class)))
    except Exception as e:
        
        raise ValueError(f"No supplier images found for search term: {search_term}, exception: {e}")

    #Contact suppliers button on first page
    random_sleep(0, 1)

    current_supplier_index = 0
    for _ in range(number_of_suppliers_to_contact):
        for j in range(current_supplier_index, len(text_elements)):
            #Close inquiry tab if any
            change_win_iter = 0
            while True:
                random_sleep(0, 1)
                try:
                    if len(driver.window_handles) > 1:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                    break
                except Exception as e:
                    if change_win_iter > max_retries:
                        print(f"Failed to switch windows after contacting supplier2, exception: {e}")
                        return supplier_set
                    change_win_iter += 1

            try:
                driver.execute_script("arguments[0].scrollIntoView();", text_elements[j])
                supplier_name = text_elements[j].text
            except Exception as e:
                print(f"Failed to get supplier name for search term: {search_term}, index: {j}, exception: {e}")
                continue
            
            random_sleep(0, 1)
            if supplier_name not in supplier_set:
                #Store supplier to avoid duplicates
                current_supplier_index = j
                supplier_set.add(supplier_name)

                #See if supplier is already in chat
                quantity = rfq_info[0]
                rfq_product_name = rfq_info[1]
                image_url = rfq_info[2]
                max_exw_price = rfq_info[3]
                max_size = rfq_info[4]
                description_tuple = (quantity, rfq_product_name, image_url, max_exw_price, max_size)
                with chat_product_lock:
                    if supplier_name in chat_product_dict:
                        current_chat_list = len(chat_product_dict[supplier_name])
                        if current_chat_list != 0:
                            #Don't contact yet if supplier is already in chat
                            # chat_product_dict[supplier_name].append(description_tuple)
                            print("image_button_break")
                            break
                    #Else statements for adding to product dict is later to ensure message is sent

                #Contact supplier
                random_sleep(1, 2)

                try:
                    driver.execute_script("arguments[0].scrollIntoView();", image_elements[j])
                    image_elements[j].click()
                except Exception as e:

                    raise ValueError(f"Failed to click supplier image for search term: {search_term}, index: {j}, exception: {e}")

                #Switch window view
                print("#Switch window view")
                random_sleep(2, 3)
                change_win_iter = 0
                while True:
                    random_sleep(1, 2)
                    try:
                        if len(driver.window_handles) > 1:
                            driver.switch_to.window(driver.window_handles[1])
                        break
                    except Exception as e:
                        if change_win_iter > max_retries:
                            
                            raise ValueError(f"Failed to switch windows after contacting supplier for search term: {search_term}, index: {j}, exception: {e}")
                        change_win_iter += 1

                #Click Contact Supplier button
                isNoSuchElement = 0
                btn_click_iter = 0
                while True:
                    random_sleep(0, 1)
                    try:
                        contact_supplier_button_xpath = "//button[@data-type-btn='inquiry']"
                        contact_button = driver.find_element(By.XPATH, contact_supplier_button_xpath)
                        contact_button.click()
                        break
                    except Exception as e:
                        if btn_click_iter >= max_retries:
                            isNoSuchElement = 1
                            break
                        time.sleep(3)
                        print("3 sec later, will retry...")
                        btn_click_iter += 1

                if isNoSuchElement == 1:
                    continue

                #Change to iframe
                popup_iter = 0
                while True:
                    random_sleep(1, 2)
                    popup_class_id = "alitalk-dialog-iframe"
                    try:
                        popup_iframe_element = driver.find_element(By.CLASS_NAME, popup_class_id)
                        driver.switch_to.frame(popup_iframe_element)
                        break
                    except Exception as e:
                        if popup_iter > max_retries:
                            
                            raise ValueError(f"Failed to switch to supplier popup for search term: {search_term}, index: {j}, exception: {e}") 
                        popup_iter += 1
                
                errors = final_input_interaction(driver, wait, description_tuple, supplier_name)
                if errors:
                    print("Error in sending message2. Continue? (y/n)")
                    pass
                    # check_error = input("Error in sending message2. Continue? (y/n)")
                    # if check_error.lower() == 'n':
                        # driver.quit()

                break

    return supplier_set

def create_chat_steps(supplier_name):
    global chat_step_dict
    global chat_step_lock
    global chat_product_dict
    global chat_product_lock

    with chat_product_lock:

        if supplier_name in chat_product_dict:
            product_tuple = chat_product_dict[supplier_name]
            product_name = ""
            quantity = 0 
            if isinstance(product_tuple, list):
                quantity = product_tuple[0][0]
                product_name = product_tuple[0][1]
            else:
                quantity = product_tuple[0]
                product_name = product_tuple[1]

            #See if chat steps already exist
            if supplier_name in chat_step_dict:
                print(f"Chat steps already exist for supplier: {supplier_name}")
                current_step_dict = chat_step_dict[supplier_name]
                return {supplier_name: current_step_dict}
            else:
                print(f"Creating new chat steps for supplier: {supplier_name}")
                #Add preset steps
                confirm = f"Are you selling {product_name}?"
                exw_price_step = f"What is the EXW price for {quantity} units?"
                # exw_price_threshold_step = f"Confirm price given is less than {exw_max_price_str}"
                sample = "Can I get a sample?"
                package_dim_step = "What are the package dimensions?"
                package_weight_step = f"What is the package weight for {quantity} units?"
                product_brand_step = "Does the product come unbranded?"
                picture = "Would I be able to get a picture?"
                spec_steps = {
                    confirm: 'unsure',
                    exw_price_step: 'unsure',
                    sample: 'unsure',
                    package_dim_step: 'unsure',
                    package_weight_step: 'unsure',
                    product_brand_step: 'unsure',
                    picture: 'unsure'
                }
                chat_step_dict[supplier_name] = spec_steps
                async_thread = threading.Thread(target=write_dict_to_file, args=(chat_step_dict_loc, chat_step_dict, chat_step_lock))
                async_thread.start()

                return {supplier_name: spec_steps}
            
        else:
            raise ValueError(f"Supplier: {supplier_name} not in chat_product_dict")
        

def resend_image(driver, supplier_name):
    global chat_product_dict
    global chat_product_lock

    #Get image url
    with chat_product_lock:
        if supplier_name in chat_product_dict:
            image_url = ""
            if isinstance(chat_product_dict[supplier_name], list):
                image_url = chat_product_dict[supplier_name][0][2] 
            else:
                image_url = chat_product_dict[supplier_name][2]
            image_download_iter = 0
            current_dir = os.path.dirname(os.path.abspath(__file__))
            tmp_file_path = os.path.join(current_dir, "tmp_image.jpg")
            while True:
                try:
                    random_sleep(1, 2)
                    image_response = requests.get(image_url, timeout=10)  # Add timeout for better control
                    if image_response.status_code == 200:
                        with open(tmp_file_path, 'wb') as tmp_file:
                            tmp_file.write(image_response.content)
                        break
                    else:
                        print(f"Error: Received status code {image_response.status_code} for URL {image_url}")
                except requests.exceptions.RequestException as e:
                    print(f"Attempt {image_download_iter + 1} failed: {e}")
                image_download_iter += 1
                if image_download_iter > max_retries:
                    print(f"Failed to download image after {max_retries} attempts")
                    return True
            #Send Image
            random_sleep(2, 4)
            try:
                js_script = "arguments[0].style.display = 'block';"
                file_input = driver.find_element(By.NAME, "file")
                driver.execute_script(js_script, file_input)
                file_input.send_keys(tmp_file_path)
            except Exception as e:
                
                raise ValueError(f"Failed to resend image, exception: {e}")

        else:
            raise ValueError(f"Supplier: {supplier_name} not in chat_product_dict")
    
def delete_chat_convo(driver, supplier_name = None):
    global chat_product_dict
    global chat_product_lock
    global chat_step_dict
    global chat_step_lock

    #Delete chat history
    if supplier_name != None:
        with chat_product_lock:
            if supplier_name in chat_product_dict:
                if isinstance(chat_product_dict[supplier_name], list):
                    chat_product_dict[supplier_name] = chat_product_dict[supplier_name].pop(0)
                else:
                    del chat_product_dict[supplier_name]
        async_thread = threading.Thread(target=write_dict_to_file, args=(chat_dict_loc, chat_product_dict, chat_product_lock))
        async_thread.start()

        with chat_step_lock:
            if supplier_name in chat_step_dict:
                del chat_step_dict[supplier_name]
        async_thread = threading.Thread(target=write_dict_to_file, args=(chat_step_dict_loc, chat_step_dict, chat_step_lock))
        async_thread.start()


    #Obtain selected element
    try:
        element_selected_css_selector = ".contact-item-container.selected"
        element_selected = driver.find_element(By.CSS_SELECTOR, element_selected_css_selector)
        element_selected.click()
    except:
        print(f"Couldn't find selected element for supplier: {supplier_name} to delete chat")
        return
    
    random_sleep(1, 2)
    try:
        open_chat_delete_option_class = ".im-next-icon.im-next-icon-ellipsis.im-next-xs"
        open_chat_delete_option = element_selected.find_element(By.CSS_SELECTOR, open_chat_delete_option_class)
        open_chat_delete_option.click()
        random_sleep(0, 1)
    except:
        raise ValueError(f"Couldn't click on open delete chat for supplier: {supplier_name}")
    
    try:
        delete_chat_xpath = "//span[@class='menu-content' and contains(text(), 'Archive')]"
        delete_chat = driver.find_element(By.XPATH, delete_chat_xpath)
        delete_chat.click()
        random_sleep(0, 1)
    except:
        print(f"Couldn't click Archive chat, so click Pin to top: {supplier_name}")
        delete_chat_xpath = "//span[@class='menu-content' and contains(text(), 'Pin to top')]"
        delete_chat = driver.find_element(By.XPATH, delete_chat_xpath)
        delete_chat.click()
        random_sleep(0, 1)
        driver.execute_script("arguments[0].scrollIntoView();", delete_chat)
        random_sleep(0, 1)
        try:
            delete_chat_xpath = "//span[@class='menu-content' and contains(text(), 'Archive')]"
            delete_chat = driver.find_element(By.XPATH, delete_chat_xpath)
            delete_chat.click()
            random_sleep(0, 1)
        except:
            raise ValueError(f"Couldn't click delete chat for supplier: {supplier_name}")
    
    random_sleep(2, 3)
    try:
        delete_chat_xpath = "//a[@class='im-next-balloon-close']"
        delete_chat = driver.find_element(By.XPATH, delete_chat_xpath)
        delete_chat.click()
        random_sleep(0, 1)
    except:
        pass

def monitor_chats(driver, wait):
    global chat_step_dict
    global chat_step_lock
    global chat_product_dict
    global chat_product_lock
    
    #Get to chat page
    chat_page_url = "https://message.alibaba.com/message/messenger.htm#/"
    driver.get(chat_page_url)
    random_sleep(5, 6)

    class RESULT:
        CHATGPT_RESPONSE_FAILED = -2
        FAILED = -1
        UNSURE = 0
        SUCCESS = 1

    success_flag = RESULT.UNSURE
    blocked_counter = 0
    print("Starting chat monitoring...")
    while True:
        print("Checking for new messages")

        print("Removing Tip_window")
        random_sleep(3, 4)

        #Check if suggestions popup
        try:
            suggestions_popup_class = "label-tip-container"
            suggestions_popup = driver.find_element(By.CLASS_NAME, suggestions_popup_class)
            random_sleep(0, 1)
   
            #Close suggestions popup
            close_suggestions_popup_xpath = ".//button[contains(@class, 'im-next-btn') and contains(@class, 'im-next-medium') and contains(@class, 'im-next-btn-primary')]"
            close_button = suggestions_popup.find_element(By.XPATH, close_suggestions_popup_xpath)
            close_button.click()
            random_sleep(0, 1)
        except:
            pass

        # Get first elements in all_tab. If not found, refresh the page
        blocked_counter = 0
        try:
            item_container_class = "contact-item-container"
            first_text_element = driver.find_element(By.CLASS_NAME, item_container_class)
            # first_text_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, item_container_class)))
            # text_elements = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, item_container_class)))
        except:
            print("Failed to select first element container")
            blocked_counter += 1
            if blocked_counter > max_retries:
                chat_page_url = "https://message.alibaba.com/message/messenger.htm#/"
                driver.get(chat_page_url)
                random_sleep(3, 4)
            continue

        print("clicking...")
        element = first_text_element
        isfirst = True
        
        while True:
            if isfirst == False:
                try:
                    parent_div = element.find_element(By.XPATH, './..')
                    next_sibling_div = parent_div.find_element(By.XPATH, 'following-sibling::div')
                    next_element = next_sibling_div.find_element(By.CLASS_NAME, item_container_class)

                    if next_element:
                        print("Clicking on the next sibling element...")
                        element = next_element
                    else:
                        print("No next sibling element found. clicking first element...")
                        element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, item_container_class)))
                except StaleElementReferenceException:
                    # Re-locate the element
                    print("StaleElementReferenceException: No next sibling element found. clicking first element...")
                    element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, item_container_class)))
                except TimeoutException:
                    print("TimeoutException: No next sibling element found. clicking first element...")
                    element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, item_container_class)))
                except NoSuchElementException:
                    print(f"NoSuchElementException: No next sibling element found. except: {e}, clicking first element...")
                    element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, item_container_class)))

            # store the next element of current one into variable and then check for success, because of deleting the current element.
            if (success_flag != RESULT.UNSURE):
                delete_chat_convo(driver, supplier_name)

            success_flag = RESULT.UNSURE
            isfirst = False
            img_src = ""
            
            random_sleep(5, 6)
            iter = 0
            while True:
                # Scroll down page
                try:
                    wait.until(EC.element_to_be_clickable(element))
                    driver.execute_script("arguments[0].scrollIntoView();", element)
                    random_sleep(0, 1)
                    element.click()
                    break
                except Exception as e:
                    print(e)
                    iter += 1
                    if(iter >= max_retries):
                        isfirst = False
                        break
                    time.sleep(5)
                    print("5 sec later, will retry...")

            if (iter >= max_retries):
                break

            random_sleep(3, 4)

            company_name_xpath = "//div[contains(@class, 'contact-item-container') and contains(@class, 'selected')]/div[@class='contact-right']/div[@class='contact-company']"
            try:
                # supplier_name = wait.until(EC.presence_of_element_located((By.XPATH, company_name_xpath))).text
                supplier_name = driver.find_element(By.XPATH, company_name_xpath).text
            
            except: 
                print("Failed to get supplier name for unread element")
                isfirst = False
                continue
                # raise ValueError(f"Couldn't get supplier name for unread element")
            
            print("xxxxxx")
            print(supplier_name)
            
            iter = 0
            while True:
                try:
                    message_item_wrapper_class = "message-item-wrapper"
                    msg_wrapper_elements = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, message_item_wrapper_class)))
                    new_message_array = []
                    for element2 in reversed(msg_wrapper_elements):
                        if 'item-right' in element2.get_attribute('class').split():
                            try:
                                child_element = element2.find_element(By.XPATH, ".//div[contains(@class, 'session-rich-content')]")
                                if (child_element):
                                    break
                            except NoSuchElementException as e:
                                print(f"Child element not found: {e}")
                            except StaleElementReferenceException as e:
                                print(f"Child element not found: {e}")
                        try:
                            # Get the text content of the 'session-rich-content' div
                            child_element = element2.find_element(By.XPATH, ".//div[contains(@class, 'session-rich-content') and contains(@class, 'text')]")
                            new_message_text = child_element.text
                            print(new_message_text)
                        except NoSuchElementException as e:
                            print(f"Child Text element not found: {e}")
                            try:
                                image_element = element2.find_element(By.XPATH, ".//div[contains(@class, 'session-rich-content') and contains(@class, 'media')]/div/img")
                                if image_element:
                                    img_src = image_element.get_attribute("src")
                            except NoSuchElementException as e:
                                print(f"Child Image element not found: {e}")
                            continue
                        new_message_array.append(new_message_text)
                    break
                except Exception as e:
                    print(e)
                    if (iter >= max_retries):
                        raise ValueError(f"Couldn't get supplier name for unread element")
                    random_sleep(2, 3)
                    iter += 1

            if len(new_message_array) == 0:
                #First message is just an image or sent you to different person (wait for another message)
                continue

            new_messages = ', '.join(new_message_array[::-1])
            print(f"New messages: {new_messages}")

            chat_step_array = []
            analysing_model = "analysing"
            max_price = 0.0
            try:
                current_step_dict = create_chat_steps(supplier_name)
                if img_src != "":
                    if (list(chat_step_dict[supplier_name].items())[6][1].lower().__contains__("unsure")):
                        print(f"img_src : {img_src}")
                        chat_step_dict[supplier_name][list(chat_step_dict[supplier_name].keys())[6]] = img_src

                for index, (key, value) in enumerate(current_step_dict[supplier_name].items(), start=1):
                    if chat_step_dict[supplier_name][key].lower().__contains__("unsure"):
                        question = f"{index}. {key}"
                        chat_step_array.append(question)

                with chat_product_lock:
                    current_product = ""
                    if isinstance(chat_product_dict[supplier_name], list):
                        current_product = chat_product_dict[supplier_name][0][1]
                        try:
                            max_price = float(re.sub(r'[^\d.]', '', chat_product_dict[supplier_name][0][3]))
                            # max_price = sanitize_price(chat_product_dict[supplier_name][0][3])
                        except:
                            raise ValueError(f"Couldn't get max price for supplier: {supplier_name} price: {chat_product_dict[supplier_name][0][3]}")
                    else:
                        current_product = chat_product_dict[supplier_name][1]
                        try:
                            max_price = float(re.sub(r'[^\d.]', '', chat_product_dict[supplier_name][3]))
                        except:
                            raise ValueError(f"Couldn't get max price for supplier: {supplier_name} price: {chat_product_dict[supplier_name][3]}")
            
            except Exception as e:
                continue

            print("confirming the chatGPT response...")
            iter = 0
            current_chat_step_string = "\n".join(chat_step_array)
            print(f"current_chat_step_string :: {current_chat_step_string}")
            
            chat_tuple = (current_chat_step_string, new_messages)
            while True:
                #Determine what questions were answered by supplier
                chat_response = query_openai(chat_tuple, analysing_model)
                print(f"Initial chat response: {chat_response}")

                if chat_response:
                    #Use regex to get indexes of answered questions use another GPT to check whether the questions were good (END if not, remove from array if good)
                    if is_json(chat_response):
                        try:
                            python_list = []
                            python_list = json.loads(chat_response)
                            print(f"[1008]: {python_list}")
                            
                            cleaned_dict = {
                                (key.strip() if isinstance(key, str) else key): (value.strip() if isinstance(value, str) else value)
                                for key, value in python_list.items()
                            }
                            print(f"[1011]: {cleaned_dict}")
                            
                            sorted_items = sorted(cleaned_dict.items(), key=lambda item: int(item[0]))
                            formatted_list = {int(key): value for key, value in sorted_items}
                            print(f"answers_dict : {formatted_list}")
                            if formatted_list == {}:
                                continue

                            break

                        except:
                            print("error")
                            continue

            print(f"[1025]: {formatted_list}")

            try:
                for index, (key, value) in enumerate(chat_step_dict[supplier_name].items(), start=1):
                    if chat_step_dict[supplier_name][key].lower().__contains__("unsure"):
                        if formatted_list[index].lower() == "no":
                            success_flag = -1
                        if (index == 2):
                            if "unsure" not in formatted_list[index].lower():
                                if (float(formatted_list[2]) > max_price):
                                    success_flag = -1

                        if (index == 7):
                            if formatted_list[index].lower().__contains__("picture"):
                                resend_image(driver, supplier_name)
                                # input("sending image...")
                            elif "unsure" in chat_step_dict[supplier_name][key].lower():
                                chat_step_dict[supplier_name][key] = formatted_list[index]
                        else:
                            chat_step_dict[supplier_name][key] = formatted_list[index]
            except Exception as e:
                print(e)

            if success_flag == RESULT.FAILED:
                print("NO___END")
                continue

            if success_flag != RESULT.CHATGPT_RESPONSE_FAILED:
                success_flag = RESULT.SUCCESS
                for (key, value) in chat_step_dict[supplier_name].items():
                    if "unsure" in chat_step_dict[supplier_name][key].lower():
                        success_flag = RESULT.UNSURE

                if (success_flag == RESULT.SUCCESS):
                    #All questions answered
                    #TODO Save info to spreadsheet
                    product_tuple = chat_product_dict[supplier_name]
                    product_name = ""
                    if isinstance(product_tuple, list):
                        product_name = product_tuple[0][1]
                    else:
                        product_name = product_tuple[1]
                    product_description_tuple = (supplier_name, list(chat_step_dict[supplier_name].items())[6][1], max_price, list(chat_step_dict[supplier_name].items())[3][1], list(chat_step_dict[supplier_name].items())[4][1])
                    googleSheet(product_name, product_description_tuple)
                    continue

            #Send questions remaining
            second_chat_question_array = []
            for index, (key, value) in enumerate(chat_step_dict[supplier_name].items(), start=1):
                if isinstance(chat_step_dict[supplier_name][key], str):
                    if chat_step_dict[supplier_name][key].lower().__contains__("unsure"):
                        print(f"value == {chat_step_dict[supplier_name][key]}")
                        question = f"{index}. {key}"
                        second_chat_question_array.append(question)
            current_chat_step_string = "\n".join(second_chat_question_array)
            print(f"[1074]:{current_chat_step_string}")

            chat_response = "Can you please answer the following questions: \n" + current_chat_step_string + "\n(preferably in 1. 2. 3, etc. Please format to avoid confusion and answer with exact numbers.)"

            message_element_class = "send-textarea"
            message_element = driver.find_element(By.CLASS_NAME, message_element_class)
            message_element.click()

            print(f"Chat response: {chat_response}")

            sentences = chat_response.split('\n')
            try:
                for sentence in sentences:
                    # If the sentence is not blank, add it to the current paragraph
                    if sentence:
                        random_sleep(0, 1)
                        message_element.send_keys(sentence)
                        random_sleep(0, 1)
                        message_element.send_keys(Keys.SHIFT, Keys.ENTER)

                input("Continue?")
                random_sleep(1, 2)
                message_element.send_keys(Keys.ENTER)
            except:
                raise ValueError(f"Couldn't send chat response for supplier: {supplier_name}")
            

def read_data():
    # Initialize session and capture data
    export_name = 'upwork_sample.xlsx'
    workbook = pyxl.load_workbook(export_name)
     
    sheet = workbook['Sheet1']

    # Initialize an empty list to store rows
    all_needed_values = []

    # Iterate over the rows in the sheet
    important_columns = {'search term','amzn link', 'order details', 'rfq quantity', 'rfq product name', 'rfq item description', 'max exw price', 'max size'}
    important_columns_indexes = set()
    for i, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        row_values = []
        error_columns = important_columns_indexes.copy()

        for j, cell in enumerate(row, start=1):
            #Get cell color
            if cell.fill.fill_type is None or cell.fill.fgColor.rgb == "FFFFFFFF":
                #Only get white cells
                value = cell.value
                if i == 1:
                    if value is not None:
                        formatted_column_name = value.strip().lower()
                        if formatted_column_name in important_columns:
                            important_columns_indexes.add(j)
                elif j in important_columns_indexes:
                    error_columns.remove(j)
                    if cell.hyperlink:
                        row_values.append(cell.hyperlink.target)
                    else:
                        row_values.append(value)

        # Add the row's values to the all_values list
        if len(row_values) != len(important_columns):
            if len(row_values) > 0:
                print(f"Error in row: {row_values}  column(s): {error_columns}")
        else:
            all_needed_values.append(row_values)
    
    return all_needed_values


def authenticate_google_sheets(json_keyfile):
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile, scope)
    client = gspread.authorize(credentials)
    return client


def set_column_headers(sheet, headers, header_format):
    sheet.append_row(headers)

    # Apply formatting to the header row
    format_cell_range(sheet, f'A1:E1', header_format)


def googleSheet(product_name, product_description_tuple):
    supplier_name, picture, price, package_demention, package_weight = product_description_tuple
    headers = ["Supplier Name", "Picture Link", "Max Price", "Package Demention", "Package Weight"]

    # Ensure the link length does not exceed 99 characters
    if len(product_name) > 99:
        sheet_name = product_name[:99]  # Cut the link to 99 characters
    else:
        sheet_name = product_name

    # Set the header format
    header_format = CellFormat(
        backgroundColor=Color(0.9, 0.9, 0.9),  # Light grey background
        textFormat=TextFormat(bold=True, fontSize=12, foregroundColor=Color(0, 0, 0)),  # Bold text
        horizontalAlignment='CENTER'
    )

    # Authenticate and get the spreadsheet
    json_keyfile = './logical-fort-420509-2592be45835c.json'
    client = authenticate_google_sheets(json_keyfile)
    spreadsheet_id = '1TOGEO2bU19X-_raSv4R2VCLPmaYMQpyqj6ps9Qb6cDY'
    spreadsheet = client.open_by_key(spreadsheet_id)

    try:
        # Check if the sheet exists, if not create it
        sheet = spreadsheet.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        # Adding a new sheet
        new_sheet = spreadsheet.add_worksheet(title=sheet_name, rows="100", cols="20")
        print(f"New sheet '{sheet_name}' added successfully.")

        # Set the headers and apply styles
        set_column_headers(new_sheet, headers, header_format)
        sheet = new_sheet

        # Get GID of the new sheet
        new_sheet_gid = new_sheet.id
        new_sheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={new_sheet_gid}"
        print(f"URL of the new sheet: {new_sheet_url}")
        add_google_sheet_link(product_name, new_sheet_url)

    # Append the new row
    new_row = [supplier_name, picture, price, package_demention, package_weight]
    sheet.append_row(new_row)
    print(f"New row added successfully: {new_row}")

def add_google_sheet_link(product_name, link):
    export_name = 'upwork_sample.xlsx'
    workbook = pyxl.load_workbook(export_name)
     
    sheet = workbook['Sheet1']

    # Get the header row to find the product name column and google sheet link column
    header_row = 1
    product_name_column_index = None
    google_sheet_link_column_index = None

    for cell in sheet[header_row]:
        if cell.value and cell.value.strip().lower() == 'rfq product name':
            product_name_column_index = cell.column
        elif cell.value and cell.value.strip().lower() == 'google sheet link':
            google_sheet_link_column_index = cell.column

    # If the google sheet link column doesn't exist, create it
    if google_sheet_link_column_index is None:
        google_sheet_link_column_index = sheet.max_column + 1
        sheet.cell(row=header_row, column=google_sheet_link_column_index, value='Google Sheet Link')

    # Ensure the product name column exists
    if product_name_column_index is None:
        print("Product name column not found.")
        return

    # Find the row with the corresponding product name and add the link in the google sheet link column
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[product_name_column_index - 1].value == product_name:
            sheet.cell(row=row[0].row, column=google_sheet_link_column_index, value=link)
            break

    with excel_lock:
        workbook.save('upwork_sample.xlsx')



def main():
    # tuple1 = ("sup_name", "url", 40, "demention", "weight")
    # googleSheet('Replacement Fliter for Vacuum Shark iz163h Replacement Filter for Vacuum Shark iz163h,2 HEPA Filters and 12 Foam Felt Kit', tuple1)
    # input("googleSheet!")

    #Load Data
    data = read_data()
    
    #Load Current Chat State
    read_chat_dicts()

    print("initialize_alibaba_search...")
    #Init Alibaba
    driver, wait = initialize_alibaba_search()

    amazon_info_list = []
    with open('amazon_info_list.pickle', 'rb') as f:
        amazon_info_list = pickle.load(f)
        print("Loaded amazon info list")
    current_len_amazon_info = len(amazon_info_list)
    print(current_len_amazon_info)

    #Send initial messages
    prompt = input("All or just monitor (all/m)? ")
    if prompt.lower() == 'all':
        first_search = False
        for i, item_row in enumerate(data[:current_len_amazon_info]):
            #Get ASIN info
            print(item_row)
            try:
                search_term_index = 0
                amazon_link_index = 1
                rfq_quantity_index = 3
                rfq_product_name_index = 4
                rfq_description_index = 5
                max_exw_price_index = 6
                max_size_index = 7
                QC_Extra_Notes_index = 9
                search_term = item_row[search_term_index]
                link_asin = ((item_row[amazon_link_index].split('"'))[1]).split('/')[-1]
                rfq_quantity = item_row[rfq_quantity_index]
                rfq_product_name = item_row[rfq_product_name_index]
                max_exw_price = item_row[max_exw_price_index]
                max_size = item_row[max_size_index]
            except:
                print(f"Error in getting data for row: {item_row}")
                continue


            main_image_url, title = amazon_info_list[i]
        
 
            print(f"[{i}] About to be Searched Title: {title}")
            #Get simplified titles
            simplified_titles = None
            title_prompt = f'Initial title: {title}'
            title_model = "titles"
            response = query_openai(title_prompt, title_model)
            if response:
                simplified_titles = response.split("\n")
                simplified_titles = [re.sub(r'^\s*[\d]+[).]\s*|^\s*-\s*', '', title).strip() for title in simplified_titles if title.strip()]
                print(f"{simplified_titles}")
            else:
                raise ValueError(f"Failed to get simplified titles for title: {title}")
    
            #Open Alibaba and send initial messages
            supplier_set = set()
            rfq_info = [rfq_quantity, rfq_product_name, main_image_url, max_exw_price, max_size]
            for j, simplified_search_term in enumerate(simplified_titles):
                print(f"-[{j}] Searching for: {simplified_search_term}")
                if i == 0 and j == 0:
                    first_search = True
                else:
                    first_search = False

                supplier_set = send_initial_message(driver, wait, simplified_search_term, rfq_info, first_search, supplier_set)


        #Monitor chats
        # monitor_chats(driver, wait)
    else:
        monitor_chats(driver, wait)

if __name__ == '__main__':
    main()


    

    